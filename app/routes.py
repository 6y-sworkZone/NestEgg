from flask import Blueprint, render_template, request, jsonify, redirect, url_for, make_response
from datetime import datetime, date, timedelta
from app import db
from app.models import *
from sqlalchemy import func, and_, or_
import json
import uuid
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

main_bp = Blueprint('main', __name__)

@main_bp.route('/')
def index():
    today = date.today()
    month_str = today.strftime('%Y-%m')
    
    accounts = Account.query.filter_by(is_archived=False).all()
    total_assets = sum(a.balance for a in accounts)
    
    income = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'income',
        func.strftime('%Y-%m', Transaction.date) == month_str
    ).scalar() or 0
    
    expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month_str
    ).scalar() or 0
    
    recent_transactions = Transaction.query.order_by(Transaction.date.desc(), Transaction.id.desc()).limit(10).all()
    
    return render_template('index.html',
        accounts=accounts,
        total_assets=total_assets,
        income=income,
        expense=expense,
        net=income - expense,
        recent_transactions=recent_transactions,
        today=today)

@main_bp.route('/accounts')
def accounts():
    accounts = Account.query.filter_by(is_archived=False).all()
    archived = Account.query.filter_by(is_archived=True).all()
    return render_template('accounts.html', accounts=accounts, archived=archived)

@main_bp.route('/api/accounts', methods=['POST'])
def create_account():
    data = request.json
    account = Account(
        name=data['name'],
        type=data['type'],
        balance=float(data['initial_balance']),
        initial_balance=float(data['initial_balance']),
        color=data.get('color', '#3498db'),
        bill_day=data.get('bill_day'),
        payment_day=data.get('payment_day'),
        credit_limit=data.get('credit_limit', 0)
    )
    db.session.add(account)
    db.session.commit()
    
    history = AccountBalanceHistory(
        account_id=account.id,
        balance=account.balance,
        record_date=date.today()
    )
    db.session.add(history)
    db.session.commit()
    
    return jsonify({'success': True, 'id': account.id})

@main_bp.route('/api/accounts/<int:id>', methods=['PUT'])
def update_account(id):
    account = Account.query.get_or_404(id)
    data = request.json
    account.name = data['name']
    account.type = data['type']
    account.color = data.get('color', account.color)
    account.bill_day = data.get('bill_day')
    account.payment_day = data.get('payment_day')
    account.credit_limit = data.get('credit_limit', 0)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/accounts/<int:id>', methods=['GET'])
def get_account(id):
    account = Account.query.get_or_404(id)
    return jsonify({
        'id': account.id,
        'name': account.name,
        'type': account.type,
        'balance': account.balance,
        'initial_balance': account.initial_balance,
        'color': account.color,
        'bill_day': account.bill_day,
        'payment_day': account.payment_day,
        'credit_limit': account.credit_limit,
        'is_archived': account.is_archived
    })

@main_bp.route('/api/accounts/<int:id>/archive', methods=['POST'])
def archive_account(id):
    account = Account.query.get_or_404(id)
    account.is_archived = not account.is_archived
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/accounts/<int:id>', methods=['DELETE'])
def delete_account(id):
    account = Account.query.get_or_404(id)
    Transaction.query.filter(
        or_(Transaction.account_id == id, Transaction.to_account_id == id)
    ).delete()
    AccountBalanceHistory.query.filter_by(account_id=id).delete()
    db.session.delete(account)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/accounts/credit-card-reminders')
def credit_card_reminders():
    today = date.today()
    accounts = Account.query.filter(
        Account.type.in_(['信用卡', 'credit']),
        Account.is_archived == False,
        Account.payment_day != None
    ).all()
    
    reminders = []
    for acc in accounts:
        due_date = date(today.year, today.month, acc.payment_day)
        if due_date < today:
            next_month = today.month % 12 + 1
            next_year = today.year + (1 if today.month == 12 else 0)
            due_date = date(next_year, next_month, acc.payment_day)
        
        days_left = (due_date - today).days
        
        if days_left <= 3:
            stmt_balance = abs(acc.balance) if acc.balance < 0 else acc.balance
            min_payment = stmt_balance
            
            reminders.append({
                'account_id': acc.id,
                'account_name': acc.name,
                'due_date': due_date.strftime('%Y-%m-%d'),
                'days_left': days_left,
                'stmt_balance': stmt_balance,
                'min_payment': min_payment
            })
    
    return jsonify(reminders)

@main_bp.route('/api/transfer', methods=['POST'])
def transfer():
    data = request.json
    from_account = Account.query.get_or_404(data['from_account_id'])
    to_account = Account.query.get_or_404(data['to_account_id'])
    amount = float(data['amount'])
    
    from_account.balance -= amount
    to_account.balance += amount
    
    trans_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    
    t1 = Transaction(
        date=trans_date,
        amount=amount,
        type='transfer',
        account_id=from_account.id,
        to_account_id=to_account.id,
        is_transfer=True,
        note=data.get('note', '')
    )
    db.session.add(t1)
    db.session.flush()
    
    t2 = Transaction(
        date=trans_date,
        amount=amount,
        type='transfer',
        account_id=to_account.id,
        to_account_id=from_account.id,
        is_transfer=True,
        transfer_pair_id=t1.id,
        note=data.get('note', '')
    )
    t1.transfer_pair_id = t2.id
    db.session.add(t2)
    
    for acc in [from_account, to_account]:
        history = AccountBalanceHistory(
            account_id=acc.id,
            balance=acc.balance,
            record_date=trans_date
        )
        db.session.add(history)
    
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/accounts/<int:id>/reconcile', methods=['POST'])
def reconcile_account(id):
    account = Account.query.get_or_404(id)
    data = request.json
    actual = float(data['actual_balance'])
    diff = actual - account.balance
    
    recon = AccountReconciliation(
        account_id=id,
        actual_balance=actual,
        system_balance=account.balance,
        difference=diff,
        notes=data.get('notes', '')
    )
    db.session.add(recon)
    
    if abs(diff) > 0.01:
        cat = Category.query.filter_by(name='其他收入').first() if diff > 0 else Category.query.filter_by(name='其他支出').first()
        t = Transaction(
            date=date.today(),
            amount=abs(diff),
            type='income' if diff > 0 else 'expense',
            category_id=cat.id if cat else None,
            account_id=id,
            note='对账调整'
        )
        db.session.add(t)
        account.balance = actual
    
    db.session.commit()
    return jsonify({'success': True, 'difference': diff})

@main_bp.route('/api/accounts/<int:id>/balance-history')
def account_balance_history(id):
    days = request.args.get('days', 30, type=int)
    start_date = date.today() - timedelta(days=days)
    
    histories = AccountBalanceHistory.query.filter(
        AccountBalanceHistory.account_id == id,
        AccountBalanceHistory.record_date >= start_date
    ).order_by(AccountBalanceHistory.record_date).all()
    
    data = [(h.record_date.strftime('%Y-%m-%d'), h.balance) for h in histories]
    return jsonify(data)

@main_bp.route('/transactions')
def transactions():
    categories = Category.query.filter_by(parent_id=None).all()
    accounts = Account.query.filter_by(is_archived=False).all()
    return render_template('transactions.html', categories=categories, accounts=accounts)

@main_bp.route('/api/transactions')
def list_transactions():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    query = Transaction.query.filter_by(is_transfer=False)
    
    if request.args.get('category_id'):
        query = query.filter_by(category_id=request.args.get('category_id', type=int))
    if request.args.get('account_id'):
        query = query.filter_by(account_id=request.args.get('account_id', type=int))
    if request.args.get('type'):
        query = query.filter_by(type=request.args.get('type'))
    if request.args.get('min_amount'):
        query = query.filter(Transaction.amount >= request.args.get('min_amount', type=float))
    if request.args.get('max_amount'):
        query = query.filter(Transaction.amount <= request.args.get('max_amount', type=float))
    if request.args.get('start_date'):
        start = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d').date()
        query = query.filter(Transaction.date >= start)
    if request.args.get('end_date'):
        end = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d').date()
        query = query.filter(Transaction.date <= end)
    if request.args.get('tag'):
        query = query.filter(Transaction.tags.like(f'%{request.args.get("tag")}%'))
    if request.args.get('keyword'):
        query = query.filter(Transaction.note.like(f'%{request.args.get("keyword")}%'))
    
    query = query.order_by(Transaction.date.desc(), Transaction.id.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'items': [{
            'id': t.id,
            'date': t.date.strftime('%Y-%m-%d'),
            'amount': t.amount,
            'type': t.type,
            'category': t.category.name if t.category else None,
            'account': t.account.name,
            'note': t.note,
            'tags': t.tags
        } for t in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages
    })

@main_bp.route('/api/transactions', methods=['POST'])
def create_transaction():
    data = request.json
    trans = Transaction(
        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
        amount=float(data['amount']),
        type=data['type'],
        category_id=data.get('category_id'),
        account_id=data['account_id'],
        note=data.get('note', ''),
        tags=data.get('tags', '')
    )
    db.session.add(trans)
    
    account = Account.query.get(data['account_id'])
    if data['type'] == 'income':
        account.balance += float(data['amount'])
    else:
        account.balance -= float(data['amount'])
    
    history = AccountBalanceHistory(
        account_id=account.id,
        balance=account.balance,
        record_date=trans.date
    )
    db.session.add(history)
    
    db.session.commit()
    return jsonify({'success': True, 'id': trans.id})

@main_bp.route('/api/transactions/<int:id>', methods=['PUT'])
def update_transaction(id):
    trans = Transaction.query.get_or_404(id)
    data = request.json
    
    old_amount = trans.amount
    old_type = trans.type
    old_account_id = trans.account_id
    
    trans.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    trans.amount = float(data['amount'])
    trans.type = data['type']
    trans.category_id = data.get('category_id')
    trans.account_id = data['account_id']
    trans.note = data.get('note', '')
    trans.tags = data.get('tags', '')
    
    if old_account_id != data['account_id']:
        old_account = Account.query.get(old_account_id)
        if old_type == 'income':
            old_account.balance -= old_amount
        else:
            old_account.balance += old_amount
    
    account = Account.query.get(data['account_id'])
    if old_account_id != data['account_id']:
        if data['type'] == 'income':
            account.balance += float(data['amount'])
        else:
            account.balance -= float(data['amount'])
    else:
        diff = float(data['amount']) - old_amount
        if old_type == data['type']:
            if data['type'] == 'income':
                account.balance += diff
            else:
                account.balance -= diff
        else:
            if data['type'] == 'income':
                account.balance += float(data['amount']) + old_amount
            else:
                account.balance -= float(data['amount']) + old_amount
    
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/transactions/<int:id>', methods=['DELETE'])
def delete_transaction(id):
    trans = Transaction.query.get_or_404(id)
    account = Account.query.get(trans.account_id)
    
    if trans.type == 'income':
        account.balance -= trans.amount
    else:
        account.balance += trans.amount
    
    db.session.delete(trans)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/categories')
def list_categories():
    type_filter = request.args.get('type')
    query = Category.query
    if type_filter:
        query = query.filter_by(type=type_filter)
    categories = query.filter_by(parent_id=None).all()
    
    def serialize(cat):
        return {
            'id': cat.id,
            'name': cat.name,
            'type': cat.type,
            'icon': cat.icon,
            'children': [serialize(c) for c in cat.children]
        }
    
    return jsonify([serialize(c) for c in categories])

@main_bp.route('/api/categories', methods=['POST'])
def create_category():
    data = request.json
    cat = Category(
        name=data['name'],
        type=data['type'],
        parent_id=data.get('parent_id'),
        icon=data.get('icon')
    )
    db.session.add(cat)
    db.session.commit()
    return jsonify({'success': True, 'id': cat.id})

@main_bp.route('/budgets')
def budgets():
    today = date.today()
    month = request.args.get('month', today.strftime('%Y-%m'))
    categories = Category.query.filter_by(type='expense').all()
    return render_template('budgets.html', month=month, categories=categories)

@main_bp.route('/api/budgets')
def list_budgets():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    budgets = Budget.query.filter_by(month=month).all()
    
    result = []
    for b in budgets:
        spent = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            Transaction.category_id == b.category_id,
            func.strftime('%Y-%m', Transaction.date) == month
        ).scalar() or 0
        
        result.append({
            'id': b.id,
            'category_id': b.category_id,
            'category_name': b.category.name,
            'amount': b.amount,
            'spent': spent,
            'percentage': min(100, (spent / b.amount * 100) if b.amount > 0 else 0)
        })
    
    return jsonify(result)

@main_bp.route('/api/budgets', methods=['POST'])
def create_budget():
    data = request.json
    Budget.query.filter_by(
        category_id=data['category_id'],
        month=data['month']
    ).delete()
    
    budget = Budget(
        category_id=data['category_id'],
        amount=float(data['amount']),
        month=data['month']
    )
    db.session.add(budget)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/budget-templates')
def list_budget_templates():
    templates = BudgetTemplate.query.all()
    return jsonify([{
        'id': t.id,
        'name': t.name,
        'items': [{'category_id': i.category_id, 'amount': i.amount} for i in t.items]
    } for t in templates])

@main_bp.route('/api/budget-templates', methods=['POST'])
def create_budget_template():
    data = request.json
    template = BudgetTemplate(name=data['name'])
    db.session.add(template)
    db.session.flush()
    
    for item in data['items']:
        db.session.add(BudgetTemplateItem(
            template_id=template.id,
            category_id=item['category_id'],
            amount=item['amount']
        ))
    
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/budgets/suggestions')
def budget_suggestions():
    today = date.today()
    
    months = []
    for i in range(3):
        d = today - timedelta(days=30 * (i + 1))
        months.append(d.strftime('%Y-%m'))
    
    categories = Category.query.filter_by(type='expense').all()
    suggestions = []
    
    for cat in categories:
        total = 0
        count = 0
        
        for month in months:
            spent = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.type == 'expense',
                Transaction.category_id == cat.id,
                func.strftime('%Y-%m', Transaction.date) == month
            ).scalar() or 0
            
            if spent > 0:
                total += spent
                count += 1
        
        avg = total / count if count > 0 else 0
        suggested = round(avg * 1.1, 2) if avg > 0 else 500
        
        if count == 0:
            category_defaults = {
                '餐饮': 1500,
                '交通': 500,
                '购物': 1000,
                '娱乐': 500,
                '居住': 3000,
                '医疗': 300,
                '教育': 500,
                '通讯': 200,
                '其他支出': 500
            }
            suggested = category_defaults.get(cat.name, 500)
        
        suggestions.append({
            'category_id': cat.id,
            'category_name': cat.name,
            'avg_3months': round(avg, 2),
            'suggested': suggested,
            'data_points': count
        })
    
    return jsonify(suggestions)

@main_bp.route('/api/budgets/annual-overview')
def budget_annual_overview():
    year = request.args.get('year', date.today().year, type=int)
    
    categories = Category.query.filter_by(type='expense').all()
    result = []
    
    for cat in categories:
        month_data = []
        for m in range(1, 13):
            month_str = f'{year}-{m:02d}'
            
            budget = Budget.query.filter_by(
                category_id=cat.id,
                month=month_str
            ).first()
            budget_amount = budget.amount if budget else 0
            
            spent = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.type == 'expense',
                Transaction.category_id == cat.id,
                func.strftime('%Y-%m', Transaction.date) == month_str
            ).scalar() or 0
            
            execution_rate = (spent / budget_amount * 100) if budget_amount > 0 else 0
            
            month_data.append({
                'month': m,
                'budget': budget_amount,
                'spent': spent,
                'execution_rate': execution_rate
            })
        
        result.append({
            'category_id': cat.id,
            'category_name': cat.name,
            'months': month_data
        })
    
    return jsonify({'year': year, 'data': result})

@main_bp.route('/investments')
def investments():
    accounts = Account.query.filter(
        or_(Account.type == '投资账户', Account.type == 'investment')
    ).all()
    return render_template('investments.html', accounts=accounts)

@main_bp.route('/api/investments')
def list_investments():
    investments = Investment.query.all()
    result = []
    for inv in investments:
        current_price = inv.current_price or inv.buy_price
        total_value = current_price * inv.quantity
        total_cost = inv.buy_price * inv.quantity
        profit = total_value - total_cost
        profit_rate = (profit / total_cost * 100) if total_cost > 0 else 0
        
        result.append({
            'id': inv.id,
            'name': inv.name,
            'code': inv.code,
            'type': inv.type,
            'buy_price': inv.buy_price,
            'quantity': inv.quantity,
            'current_price': current_price,
            'total_value': total_value,
            'profit': profit,
            'profit_rate': profit_rate,
            'account_name': inv.account.name
        })
    
    return jsonify(result)

@main_bp.route('/api/investments', methods=['POST'])
def create_investment():
    data = request.json
    inv = Investment(
        account_id=data['account_id'],
        name=data['name'],
        code=data.get('code', ''),
        type=data.get('type', ''),
        buy_price=float(data['buy_price']),
        quantity=float(data['quantity']),
        current_price=float(data['buy_price']),
        buy_date=datetime.strptime(data['buy_date'], '%Y-%m-%d').date(),
        notes=data.get('notes', '')
    )
    db.session.add(inv)
    db.session.commit()
    return jsonify({'success': True, 'id': inv.id})

@main_bp.route('/api/investments/<int:id>/update-price', methods=['POST'])
def update_investment_price(id):
    inv = Investment.query.get_or_404(id)
    inv.current_price = float(request.json['current_price'])
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/investments/<int:id>', methods=['DELETE'])
def delete_investment(id):
    inv = Investment.query.get_or_404(id)
    db.session.delete(inv)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/investments/profit-trend')
def investment_profit_trend():
    months = db.session.query(
        func.strftime('%Y-%m', Transaction.date).label('month'),
    ).join(Dividend, Transaction.account_id == Dividend.investment_id, isouter=True).filter(
        or_(Transaction.type == 'income', Dividend.id != None)
    ).group_by(func.strftime('%Y-%m', Transaction.date)).order_by(
        func.strftime('%Y-%m', Transaction.date).desc()
    ).limit(12).all()
    
    result = []
    for month, in reversed(months):
        month_start = datetime.strptime(month + '-01', '%Y-%m-%d').date()
        if month_start.month == 12:
            next_month = date(month_start.year + 1, 1, 1)
        else:
            next_month = date(month_start.year, month_start.month + 1, 1)
        
        investments = Investment.query.all()
        total_profit = 0
        
        for inv in investments:
            if inv.buy_date < next_month:
                current_price = inv.current_price or inv.buy_price
                profit = (current_price - inv.buy_price) * inv.quantity
                total_profit += profit
        
        dividends = Dividend.query.filter(
            func.strftime('%Y-%m', Dividend.date) == month
        ).all()
        for div in dividends:
            total_profit += div.amount
        
        result.append({'month': month, 'profit': round(total_profit, 2)})
    
    if not result:
        today = date.today()
        for i in range(5, -1, -1):
            m = today - timedelta(days=30 * i)
            result.append({'month': m.strftime('%Y-%m'), 'profit': 0})
    
    return jsonify(result)

@main_bp.route('/api/investments/asset-allocation')
def asset_allocation():
    investments = Investment.query.all()
    accounts = Account.query.filter_by(is_archived=False).all()
    
    cash_total = sum(a.balance for a in accounts if a.type not in ['投资账户', 'investment'])
    
    type_values = {'股票': 0, '基金': 0, '定期': 0, '债券': 0, '其他': 0}
    
    for inv in investments:
        current_price = inv.current_price or inv.buy_price
        value = current_price * inv.quantity
        inv_type = inv.type or '其他'
        if inv_type in type_values:
            type_values[inv_type] += value
        else:
            type_values['其他'] += value
    
    total_assets = cash_total + sum(type_values.values())
    
    if total_assets == 0:
        return jsonify({
            'cash': 0, 'stocks': 0, 'funds': 0, 'fixed': 0, 
            'bonds': 0, 'other': 0, 'total': 0, 'suggestions': ['暂无资产，开始投资吧！']
        })
    
    suggestions = []
    cash_ratio = cash_total / total_assets * 100
    stock_ratio = type_values['股票'] / total_assets * 100
    fund_ratio = type_values['基金'] / total_assets * 100
    
    if cash_ratio > 50:
        suggestions.append('现金占比过高(%.1f%%)，建议适当投资' % cash_ratio)
    elif cash_ratio < 10:
        suggestions.append('现金占比过低(%.1f%%)，建议保留应急资金' % cash_ratio)
    
    if stock_ratio > 60:
        suggestions.append('股票占比过高(%.1f%%)，建议分散风险' % stock_ratio)
    
    if fund_ratio < 20 and total_assets > 10000:
        suggestions.append('建议配置基金以分散投资风险')
    
    if stock_ratio + fund_ratio < 30 and total_assets > 10000:
        suggestions.append('权益类资产占比偏低，可适当增加配置')
    
    return jsonify({
        'cash': round(cash_total, 2),
        'stocks': round(type_values['股票'], 2),
        'funds': round(type_values['基金'], 2),
        'fixed': round(type_values['定期'], 2),
        'bonds': round(type_values['债券'], 2),
        'other': round(type_values['其他'], 2),
        'total': round(total_assets, 2),
        'cash_ratio': round(cash_ratio, 1),
        'stock_ratio': round(stock_ratio, 1),
        'fund_ratio': round(fund_ratio, 1),
        'suggestions': suggestions if suggestions else ['资产配置合理，继续保持！']
    })

@main_bp.route('/api/investments/<int:id>/dividends')
def list_dividends(id):
    dividends = Dividend.query.filter_by(investment_id=id).order_by(Dividend.date.desc()).all()
    return jsonify([{
        'id': d.id,
        'amount': d.amount,
        'date': d.date.strftime('%Y-%m-%d'),
        'note': d.note
    } for d in dividends])

@main_bp.route('/api/investments/<int:id>/dividends', methods=['POST'])
def add_dividend(id):
    Investment.query.get_or_404(id)
    data = request.json
    div = Dividend(
        investment_id=id,
        amount=float(data['amount']),
        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
        note=data.get('note', '')
    )
    db.session.add(div)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/investments/<int:id>/notes')
def list_investment_notes(id):
    notes = InvestmentNote.query.filter_by(investment_id=id).order_by(InvestmentNote.created_at.desc()).all()
    return jsonify([{
        'id': n.id,
        'content': n.content,
        'created_at': n.created_at.strftime('%Y-%m-%d %H:%M')
    } for n in notes])

@main_bp.route('/api/investments/<int:id>/notes', methods=['POST'])
def add_investment_note(id):
    Investment.query.get_or_404(id)
    data = request.json
    note = InvestmentNote(
        investment_id=id,
        content=data['content']
    )
    db.session.add(note)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/savings')
def savings():
    return render_template('savings.html')

@main_bp.route('/api/savings')
def list_savings_goals():
    goals = SavingsGoal.query.filter_by(is_completed=False).all()
    completed = SavingsGoal.query.filter_by(is_completed=True).all()
    
    def serialize(goal):
        progress = (goal.current_amount / goal.target_amount * 100) if goal.target_amount > 0 else 0
        monthly_needed = 0
        if goal.deadline and not goal.is_completed:
            days_left = (goal.deadline - date.today()).days
            months_left = max(1, days_left / 30)
            remaining = goal.target_amount - goal.current_amount
            monthly_needed = remaining / months_left
        
        return {
            'id': goal.id,
            'name': goal.name,
            'target_amount': goal.target_amount,
            'current_amount': goal.current_amount,
            'progress': progress,
            'deadline': goal.deadline.strftime('%Y-%m-%d') if goal.deadline else None,
            'priority': goal.priority,
            'monthly_needed': monthly_needed,
            'deposits': [{'amount': d.amount, 'date': d.date.strftime('%Y-%m-%d'), 'note': d.note} for d in goal.deposits]
        }
    
    return jsonify({
        'active': [serialize(g) for g in goals],
        'completed': [serialize(g) for g in completed]
    })

@main_bp.route('/api/savings', methods=['POST'])
def create_savings_goal():
    data = request.json
    goal = SavingsGoal(
        name=data['name'],
        target_amount=float(data['target_amount']),
        current_amount=float(data.get('current_amount', 0)),
        deadline=datetime.strptime(data['deadline'], '%Y-%m-%d').date() if data.get('deadline') else None,
        account_id=data.get('account_id'),
        priority=data.get('priority', 1)
    )
    db.session.add(goal)
    db.session.commit()
    return jsonify({'success': True, 'id': goal.id})

@main_bp.route('/api/savings/<int:id>/deposit', methods=['POST'])
def deposit_to_goal(id):
    goal = SavingsGoal.query.get_or_404(id)
    data = request.json
    amount = float(data['amount'])
    
    goal.current_amount += amount
    
    deposit = GoalDeposit(
        goal_id=id,
        amount=amount,
        note=data.get('note', '')
    )
    db.session.add(deposit)
    
    if goal.current_amount >= goal.target_amount:
        goal.is_completed = True
    
    db.session.commit()
    return jsonify({'success': True, 'completed': goal.is_completed})

@main_bp.route('/api/savings/<int:id>', methods=['PUT'])
def update_savings_goal(id):
    goal = SavingsGoal.query.get_or_404(id)
    data = request.json
    goal.name = data['name']
    goal.target_amount = float(data['target_amount'])
    goal.deadline = datetime.strptime(data['deadline'], '%Y-%m-%d').date() if data.get('deadline') else None
    goal.priority = data.get('priority', 1)
    goal.is_completed = False
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/savings/<int:id>', methods=['DELETE'])
def delete_savings_goal(id):
    goal = SavingsGoal.query.get_or_404(id)
    GoalDeposit.query.filter_by(goal_id=id).delete()
    db.session.delete(goal)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/analysis')
def analysis():
    return render_template('analysis.html')

@main_bp.route('/api/analysis/monthly-summary')
def monthly_summary():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    income = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'income',
        func.strftime('%Y-%m', Transaction.date) == month
    ).scalar() or 0
    
    expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).scalar() or 0
    
    by_category = db.session.query(
        Category.name,
        func.sum(Transaction.amount)
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).group_by(Category.id).all()
    
    return jsonify({
        'income': income,
        'expense': expense,
        'net': income - expense,
        'by_category': [{'name': c, 'amount': a} for c, a in by_category]
    })

@main_bp.route('/api/analysis/daily-trend')
def daily_trend():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    days = db.session.query(
        func.date(Transaction.date),
        func.sum(Transaction.amount)
    ).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).group_by(func.date(Transaction.date)).order_by(func.date(Transaction.date)).all()
    
    return jsonify([{'date': d, 'amount': a} for d, a in days])

@main_bp.route('/api/analysis/weekday-vs-weekend')
def weekday_vs_weekend():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    weekday_expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month,
        func.strftime('%w', Transaction.date).between('1', '5')
    ).scalar() or 0
    
    weekend_expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month,
        func.strftime('%w', Transaction.date).in_(['0', '6'])
    ).scalar() or 0
    
    return jsonify({
        'weekday': weekday_expense,
        'weekend': weekend_expense
    })

@main_bp.route('/api/analysis/anomaly-detection')
def anomaly_detection():
    today = date.today()
    month = request.args.get('month', today.strftime('%Y-%m'))
    
    first_day = today.replace(day=1)
    last_month = (first_day - timedelta(days=1)).strftime('%Y-%m')
    
    current_cats = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).group_by(Category.id).all()
    
    last_cats = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == last_month
    ).group_by(Category.id).all()
    
    last_dict = {name: total for name, total in last_cats}
    
    anomalies = []
    for name, total in current_cats:
        if name in last_dict and last_dict[name] > 0:
            change_rate = (total - last_dict[name]) / last_dict[name] * 100
            if change_rate >= 50:
                anomalies.append({
                    'category': name,
                    'current': total,
                    'last': last_dict[name],
                    'change_rate': change_rate
                })
    
    return jsonify(anomalies)

@main_bp.route('/api/analysis/trend-prediction')
def trend_prediction():
    months_data = db.session.query(
        func.strftime('%Y-%m', Transaction.date).label('month'),
        func.sum(Transaction.amount).label('total')
    ).filter(
        Transaction.type == 'expense'
    ).group_by(func.strftime('%Y-%m', Transaction.date)).order_by(
        func.strftime('%Y-%m', Transaction.date)
    ).limit(6).all()
    
    if len(months_data) < 2:
        return jsonify({'predicted': 0, 'months': [], 'amounts': []})
    
    months = [m for m, _ in months_data]
    amounts = [a for _, a in months_data]
    
    n = len(months)
    x = list(range(n))
    sum_x = sum(x)
    sum_y = sum(amounts)
    sum_xy = sum(x[i] * amounts[i] for i in range(n))
    sum_x2 = sum(xi ** 2 for xi in x)
    
    slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2) if (n * sum_x2 - sum_x ** 2) != 0 else 0
    intercept = (sum_y - slope * sum_x) / n
    
    predicted = max(0, slope * n + intercept)
    
    return jsonify({
        'predicted': round(predicted, 2),
        'months': months,
        'amounts': amounts
    })

@main_bp.route('/api/analysis/yoy-mom')
def yoy_mom_analysis():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    y, m = map(int, month.split('-'))
    
    current = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).scalar() or 0
    
    last_month = date(y, m, 1) - timedelta(days=1)
    mom_month = last_month.strftime('%Y-%m')
    mom_value = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == mom_month
    ).scalar() or 0
    
    yoy_year = y - 1
    yoy_month = f'{yoy_year}-{m:02d}'
    yoy_value = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == yoy_month
    ).scalar() or 0
    
    mom_rate = ((current - mom_value) / mom_value * 100) if mom_value > 0 else 0
    yoy_rate = ((current - yoy_value) / yoy_value * 100) if yoy_value > 0 else 0
    
    return jsonify({
        'current': round(current, 2),
        'mom': round(mom_value, 2),
        'mom_rate': round(mom_rate, 1),
        'yoy': round(yoy_value, 2),
        'yoy_rate': round(yoy_rate, 1),
        'mom_month': mom_month,
        'yoy_month': yoy_month
    })

@main_bp.route('/api/analysis/annual-summary')
def annual_summary():
    year = request.args.get('year', date.today().year, type=int)
    
    income = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'income',
        func.strftime('%Y', Transaction.date) == str(year)
    ).scalar() or 0
    
    expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y', Transaction.date) == str(year)
    ).scalar() or 0
    
    net_savings = income - expense
    
    top_category = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        func.strftime('%Y', Transaction.date) == str(year)
    ).group_by(Category.id).order_by(func.sum(Transaction.amount).desc()).first()
    
    months_data = []
    for m in range(1, 13):
        month_str = f'{year}-{m:02d}'
        m_income = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.type == 'income',
            func.strftime('%Y-%m', Transaction.date) == month_str
        ).scalar() or 0
        m_expense = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            func.strftime('%Y-%m', Transaction.date) == month_str
        ).scalar() or 0
        months_data.append({'month': m, 'income': m_income, 'expense': m_expense, 'savings': m_income - m_expense})
    
    savings_ranked = sorted(months_data, key=lambda x: x['savings'], reverse=True)
    best_month = savings_ranked[0]['month'] if savings_ranked else 0
    
    return jsonify({
        'year': year,
        'total_income': round(income, 2),
        'total_expense': round(expense, 2),
        'net_savings': round(net_savings, 2),
        'savings_rate': round(net_savings / income * 100, 1) if income > 0 else 0,
        'top_expense_category': top_category.name if top_category else '无',
        'top_expense_amount': round(top_category.total, 2) if top_category else 0,
        'best_savings_month': best_month,
        'months': months_data
    })

@main_bp.route('/reports')
def reports():
    accounts = Account.query.filter_by(is_archived=False).all()
    return render_template('reports.html', accounts=accounts)

@main_bp.route('/api/reports/export/excel')
def export_excel():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "财务报表"
    
    headers = ['日期', '类型', '分类', '账户', '金额', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    transactions = Transaction.query.filter(
        func.strftime('%Y-%m', Transaction.date) == month
    ).order_by(Transaction.date).all()
    
    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value='收入' if t.type == 'income' else '支出')
        ws.cell(row=row, column=3, value=t.category.name if t.category else '')
        ws.cell(row=row, column=4, value=t.account.name)
        ws.cell(row=row, column=5, value=t.amount)
        ws.cell(row=row, column=6, value=t.note)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=report_{month}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@main_bp.route('/api/reports/export/pdf')
def export_pdf():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    transactions = Transaction.query.filter(
        func.strftime('%Y-%m', Transaction.date) == month
    ).order_by(Transaction.date).all()
    
    income = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'income',
        func.strftime('%Y-%m', Transaction.date) == month
    ).scalar() or 0
    
    expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).scalar() or 0
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    elements.append(Paragraph(f'{month} 月度财务报表', styles['Title']))
    elements.append(Spacer(1, 20))
    
    summary_data = [
        ['项目', '金额'],
        ['总收入', f'¥{income:,.2f}'],
        ['总支出', f'¥{expense:,.2f}'],
        ['净收入', f'¥{income - expense:,.2f}']
    ]
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph('收支明细', styles['Heading2']))
    
    trans_data = [['日期', '类型', '分类', '账户', '金额', '备注']]
    for t in transactions:
        trans_data.append([
            t.date.strftime('%Y-%m-%d'),
            '收入' if t.type == 'income' else '支出',
            t.category.name if t.category else '',
            t.account.name,
            f'¥{t.amount:,.2f}',
            t.note or ''
        ])
    
    trans_table = Table(trans_data)
    trans_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
    ]))
    elements.append(trans_table)
    
    doc.build(elements)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=report_{month}.pdf"
    response.headers["Content-type"] = "application/pdf"
    return response

@main_bp.route('/api/health-score')
def health_score():
    today = date.today()
    month = today.strftime('%Y-%m')
    last_month = (today.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
    
    income = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'income',
        func.strftime('%Y-%m', Transaction.date) == month
    ).scalar() or 0
    
    expense = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).scalar() or 0
    
    savings_rate = (income - expense) / income * 100 if income > 0 else 0
    savings_score = min(100, savings_rate * 2)
    
    budgets = Budget.query.filter_by(month=month).all()
    budget_count = len(budgets)
    budget_ok_count = 0
    for b in budgets:
        spent = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            Transaction.category_id == b.category_id,
            func.strftime('%Y-%m', Transaction.date) == month
        ).scalar() or 0
        if spent <= b.amount:
            budget_ok_count += 1
    
    budget_score = (budget_ok_count / budget_count * 100) if budget_count > 0 else 100
    
    investments = Investment.query.all()
    inv_types = set(inv.type for inv in investments if inv.type)
    diversity_score = min(100, len(inv_types) * 25)
    
    total_score = int((savings_score * 0.4 + budget_score * 0.3 + diversity_score * 0.3))
    
    suggestions = []
    if savings_rate < 20:
        suggestions.append("建议提高储蓄率，目标至少20%")
    
    cats = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        func.strftime('%Y-%m', Transaction.date) == month
    ).group_by(Category.id).order_by(func.sum(Transaction.amount).desc()).limit(3).all()
    
    for name, amount in cats:
        if amount and amount > 500:
            suggestions.append(f"你每月{name}支出 {amount:.0f}元，一年就是 {amount*12:.0f} 元")
    
    return jsonify({
        'score': total_score,
        'savings_rate': savings_rate,
        'savings_score': savings_score,
        'budget_score': budget_score,
        'diversity_score': diversity_score,
        'suggestions': suggestions
    })

@main_bp.route('/api/recurring')
def list_recurring():
    recurring = RecurringTransaction.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': r.id,
        'name': r.name,
        'amount': r.amount,
        'type': r.type,
        'frequency': r.frequency,
        'day_of_month': r.day_of_month,
        'next_run_date': r.next_run_date.strftime('%Y-%m-%d')
    } for r in recurring])

@main_bp.route('/api/recurring', methods=['POST'])
def create_recurring():
    data = request.json
    r = RecurringTransaction(
        name=data['name'],
        amount=float(data['amount']),
        type=data['type'],
        category_id=data.get('category_id'),
        account_id=data['account_id'],
        frequency=data.get('frequency', 'monthly'),
        day_of_month=data.get('day_of_month', 1)
    )
    db.session.add(r)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/recurring/<int:id>', methods=['DELETE'])
def delete_recurring(id):
    r = RecurringTransaction.query.get_or_404(id)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'success': True})

@main_bp.route('/api/reports/share', methods=['POST'])
def create_report_share():
    data = request.form
    share_code = uuid.uuid4().hex[:16]
    
    expires_at = datetime.utcnow() + timedelta(days=7)
    
    start_date = None
    end_date = None
    
    if data.get('start_date') and data.get('end_date'):
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
    elif data.get('month'):
        year, month = map(int, data['month'].split('-'))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    share = ReportShare(
        share_code=share_code,
        report_type='custom',
        start_date=start_date,
        end_date=end_date,
        include_transactions=data.get('include_transactions', 'true') == 'true',
        include_summary=data.get('include_summary', 'true') == 'true',
        include_budget=data.get('include_budget', 'true') == 'true',
        expires_at=expires_at
    )
    db.session.add(share)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'share_code': share_code,
        'share_url': f'/share/{share_code}'
    })

@main_bp.route('/share/<code>')
def share_view(code):
    share = ReportShare.query.filter_by(share_code=code).first()
    
    if not share:
        return render_template('share.html', share_code=code, error='分享链接不存在')
    
    if share.expires_at and datetime.utcnow() > share.expires_at:
        return render_template('share.html', share_code=code, error='分享链接已过期')
    
    start_date = share.start_date or date.today().replace(day=1)
    end_date = share.end_date or date.today()
    
    include_transactions = share.include_transactions if hasattr(share, 'include_transactions') else True
    
    return render_template('share.html',
        share_code=code,
        start_date=start_date.strftime('%Y-%m-%d'),
        end_date=end_date.strftime('%Y-%m-%d'),
        expires_at=share.expires_at.strftime('%Y-%m-%d %H:%M') if share.expires_at else '无',
        include_transactions=include_transactions,
        error=None)

@main_bp.route('/api/reports/share/<code>/data')
def get_share_data(code):
    share = ReportShare.query.filter_by(share_code=code).first()
    
    if not share or (share.expires_at and datetime.utcnow() > share.expires_at):
        return jsonify({'error': '分享链接无效或已过期'}), 404
    
    start_date = share.start_date or date.today().replace(day=1)
    end_date = share.end_date or date.today()
    
    transactions = Transaction.query.filter(
        Transaction.date >= start_date,
        Transaction.date <= end_date
    ).order_by(Transaction.date.desc()).all()
    
    income = sum(t.amount for t in transactions if t.type == 'income')
    expense = sum(t.amount for t in transactions if t.type == 'expense')
    
    expense_by_category = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('amount')
    ).join(Transaction, Transaction.category_id == Category.id).filter(
        Transaction.date >= start_date,
        Transaction.date <= end_date,
        Transaction.type == 'expense'
    ).group_by(Category.name).all()
    
    return jsonify({
        'income': income,
        'expense': expense,
        'by_category': [{'name': c[0], 'amount': c[1]} for c in expense_by_category],
        'transactions': [{
            'date': t.date.strftime('%Y-%m-%d'),
            'category': t.category.name if t.category else '-',
            'description': t.description,
            'amount': t.amount,
            'type': t.type
        } for t in transactions[:50]] if share.include_transactions else []
    })
